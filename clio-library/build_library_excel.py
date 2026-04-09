#!/usr/bin/env python3
"""
Arvas Familjebibliotek — Excel-mall v1.0
Bygger en återanvändbar Excel-template som läser data från en CSV-fil.
"""

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.styles.numbers import FORMAT_TEXT
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation
import datetime

# ── AIAB Färgpalett ──────────────────────────────────────────
C_CREAM       = "F7F2E8"
C_CREAM_DARK  = "EDE5D0"
C_INPUT_BG    = "FDF6DC"
C_INPUT_BORD  = "C8A84B"
C_HEADER_DARK = "2A3F6F"
C_HEADER_MED  = "4A6FA5"
C_GOLD        = "C8A84B"
C_GOLD_LIGHT  = "E8C96A"
C_INK         = "3D2E0A"
C_INK_SOFT    = "7A6A4A"
C_FORMULA     = "2A3F6F"
C_INPUT_VAL   = "B8860B"
C_POSITIVE    = "27622A"
C_NEGATIVE    = "B83232"
C_WHITE       = "FFFFFF"
C_SHELF       = "6B4A2A"

def fill(hex_color):
    return PatternFill("solid", start_color=hex_color, end_color=hex_color)

def font(bold=False, italic=False, color=C_INK, size=10, name="Arial"):
    return Font(name=name, bold=bold, italic=italic, color=color, size=size)

def border_thin(sides="all"):
    s = Side(style="thin", color="C8A84B")
    n = Side(style=None)
    t = s if "all" in sides or "t" in sides else n
    b = s if "all" in sides or "b" in sides else n
    l = s if "all" in sides or "l" in sides else n
    r = s if "all" in sides or "r" in sides else n
    return Border(top=t, bottom=b, left=l, right=r)

def border_bottom():
    s = Side(style="thin", color="C8A84B")
    return Border(bottom=s)

wb = Workbook()

# ════════════════════════════════════════════════════════════
# FLIK 1 – Instructions
# ════════════════════════════════════════════════════════════
ws_inst = wb.active
ws_inst.title = "Instructions"
ws_inst.sheet_properties.tabColor = C_HEADER_DARK

ws_inst.column_dimensions["A"].width = 3
ws_inst.column_dimensions["B"].width = 28
ws_inst.column_dimensions["C"].width = 65

# Header
ws_inst.merge_cells("A1:C1")
ws_inst["A1"] = "📚 Arvas Familjebibliotek — Excel-mall v1.0"
ws_inst["A1"].font = Font(name="Arial", bold=True, size=16, color=C_GOLD_LIGHT)
ws_inst["A1"].fill = fill(C_HEADER_DARK)
ws_inst["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws_inst.row_dimensions[1].height = 36

ws_inst.merge_cells("A2:C2")
ws_inst["A2"] = "Mall för export och analys av familjebiblioteket. Källdata: CSV exporterad från Notion."
ws_inst["A2"].font = Font(name="Arial", italic=True, size=10, color=C_INK_SOFT)
ws_inst["A2"].fill = fill(C_CREAM_DARK)
ws_inst["A2"].alignment = Alignment(horizontal="center", vertical="center")
ws_inst.row_dimensions[2].height = 22

# Sections
sections = [
    ("", "", ""),
    ("SYFTE", "", ""),
    ("", "Vad är det här?", "En återanvändbar Excel-mall för Arvas Familjebibliotek. Filen innehåller ingen bokdata — all data hämtas från en CSV-exporterad från Notion."),
    ("", "Primära användningsfall", "• Filtrera och sortera bokhyllan per ägare, hylla eller genre\n• Ulrikas betygsexport med Goodreads-länkar\n• Prissättning inför Bokbörsens\n• Periodisk snapshot (~en gång i månaden)"),
    ("", "", ""),
    ("KOM IGÅNG", "", ""),
    ("", "Steg 1 — Exportera CSV", "Kör export_to_csv.py från projektmappen:\n  python export_to_csv.py --owner FrUlle\nCSV-filen sparas som arvas_bibliotek.csv i samma mapp."),
    ("", "Steg 2 — Importera data", "Gå till fliken [Data] och klicka på knappen 'Uppdatera från CSV'\nEller: Kopiera innehållet från CSV och klistra in från rad 2 i Data-fliken."),
    ("", "Steg 3 — Välj ägare", "Gå till fliken [Settings] och välj vems böcker som ska visas.\nAnalysis-fliken uppdateras automatiskt."),
    ("", "", ""),
    ("FLIKAR", "", ""),
    ("", "Instructions", "Den här sidan. Syfte, steg-för-steg och begränsningar."),
    ("", "Settings", "Justerbara parametrar: ägare, filter, exportdatum."),
    ("", "Data", "Rådata importerad från CSV. En rad per bok. Redigera ej manuellt."),
    ("", "Analysis", "Pivottabeller, statistik och diagram baserade på Data + Settings."),
    ("", "File History", "Versionslogg för filen."),
    ("", "", ""),
    ("BEGRÄNSNINGAR", "", ""),
    ("", "Ingen live-koppling", "Excel hämtar INTE data automatiskt. Kör export-scriptet och importera manuellt."),
    ("", "Ägare-filter", "Filtreringen sker i Settings. Analysfliken visar alltid vald ägares böcker."),
    ("", "Formler", "Redigera inte formlerna i Analysis-fliken utan att förstå referenserna."),
]

row = 3
for s in sections:
    label, key, val = s
    if label and not key:
        ws_inst.merge_cells(f"A{row}:C{row}")
        ws_inst[f"A{row}"] = f"  {label}"
        ws_inst[f"A{row}"].font = Font(name="Arial", bold=True, size=10, color=C_WHITE)
        ws_inst[f"A{row}"].fill = fill(C_HEADER_MED)
        ws_inst.row_dimensions[row].height = 20
    elif key:
        ws_inst[f"B{row}"] = key
        ws_inst[f"B{row}"].font = Font(name="Arial", bold=True, size=9, color=C_INK)
        ws_inst[f"B{row}"].fill = fill(C_CREAM)
        ws_inst[f"C{row}"] = val
        ws_inst[f"C{row}"].font = Font(name="Arial", size=9, color=C_INK_SOFT)
        ws_inst[f"C{row}"].fill = fill(C_CREAM)
        ws_inst[f"C{row}"].alignment = Alignment(wrap_text=True)
        if "\n" in val:
            ws_inst.row_dimensions[row].height = 14 * (val.count("\n") + 1)
        else:
            ws_inst.row_dimensions[row].height = 18
    else:
        ws_inst.row_dimensions[row].height = 8
    row += 1

# ════════════════════════════════════════════════════════════
# FLIK 2 – Settings
# ════════════════════════════════════════════════════════════
ws_set = wb.create_sheet("Settings")
ws_set.sheet_properties.tabColor = C_GOLD

ws_set.column_dimensions["A"].width = 3
ws_set.column_dimensions["B"].width = 28
ws_set.column_dimensions["C"].width = 25
ws_set.column_dimensions["D"].width = 40

ws_set.merge_cells("A1:D1")
ws_set["A1"] = "⚙️ Inställningar"
ws_set["A1"].font = Font(name="Arial", bold=True, size=14, color=C_WHITE)
ws_set["A1"].fill = fill(C_HEADER_DARK)
ws_set["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws_set.row_dimensions[1].height = 32

ws_set.merge_cells("A2:D2")
ws_set["A2"] = "Gula celler = inmatningsfält. Ändra dessa för att styra Analysis-fliken."
ws_set["A2"].font = Font(name="Arial", italic=True, size=9, color=C_INK_SOFT)
ws_set["A2"].fill = fill(C_CREAM_DARK)
ws_set["A2"].alignment = Alignment(horizontal="center")
ws_set.row_dimensions[2].height = 18

# Settings header row
for col, txt in [(2, "Parameter"), (3, "Värde"), (4, "Förklaring")]:
    c = ws_set.cell(row=3, column=col, value=txt)
    c.font = Font(name="Arial", bold=True, size=9, color=C_WHITE)
    c.fill = fill(C_HEADER_MED)
    c.alignment = Alignment(horizontal="center")
    c.border = border_thin()
ws_set.row_dimensions[3].height = 20

settings_rows = [
    ("Hyllägare", "FrUlle", "Filtrera böcker per ägare. FrUlle = Fredrik + Ulrika. Ändra till enskild person för deras böcker."),
    ("Exportdatum", datetime.date.today().strftime("%Y-%m-%d"), "Datum för senaste CSV-export från Notion."),
    ("Visa elektroniska", "Ja", "Ja/Nej — inkludera e-böcker och ljudböcker i listan."),
    ("Visa okänd författare", "Ja", "Ja/Nej — inkludera böcker med okänd författare."),
    ("Sortera efter", "Hyllplats", "Hyllplats / Titel / Författare / Språk"),
    ("Exportnamn", "arvas_bibliotek", "Basnamn för export. Datum läggs till automatiskt."),
]

for i, (param, val, expl) in enumerate(settings_rows):
    r = 4 + i
    ws_set[f"B{r}"] = param
    ws_set[f"B{r}"].font = Font(name="Arial", bold=True, size=9, color=C_INK)
    ws_set[f"B{r}"].fill = fill(C_CREAM)
    ws_set[f"B{r}"].border = border_thin()

    ws_set[f"C{r}"] = val
    ws_set[f"C{r}"].font = Font(name="Arial", bold=True, size=9, color=C_INPUT_VAL)
    ws_set[f"C{r}"].fill = fill(C_INPUT_BG)
    ws_set[f"C{r}"].border = Border(
        top=Side(style="thin", color=C_INPUT_BORD),
        bottom=Side(style="thin", color=C_INPUT_BORD),
        left=Side(style="thin", color=C_INPUT_BORD),
        right=Side(style="thin", color=C_INPUT_BORD),
    )
    ws_set[f"C{r}"].alignment = Alignment(horizontal="center")

    ws_set[f"D{r}"] = expl
    ws_set[f"D{r}"].font = Font(name="Arial", italic=True, size=9, color=C_INK_SOFT)
    ws_set[f"D{r}"].fill = fill(C_CREAM)
    ws_set[f"D{r}"].border = border_thin()
    ws_set[f"D{r}"].alignment = Alignment(wrap_text=True)
    ws_set.row_dimensions[r].height = 20

# Named ranges for key settings (referenced from Analysis)
# C4 = Hyllägare, C5 = Exportdatum, etc.

# ════════════════════════════════════════════════════════════
# FLIK 3 – Data
# ════════════════════════════════════════════════════════════
ws_data = wb.create_sheet("Data")
ws_data.sheet_properties.tabColor = C_HEADER_MED

ws_data.column_dimensions["A"].width = 5   # #
ws_data.column_dimensions["B"].width = 42  # Titel
ws_data.column_dimensions["C"].width = 28  # Författare
ws_data.column_dimensions["D"].width = 10  # Hyllplats
ws_data.column_dimensions["E"].width = 8   # Språk
ws_data.column_dimensions["F"].width = 12  # Ägare
ws_data.column_dimensions["G"].width = 8   # År
ws_data.column_dimensions["H"].width = 18  # ISBN
ws_data.column_dimensions["I"].width = 22  # Förlag
ws_data.column_dimensions["J"].width = 14  # Format
ws_data.column_dimensions["K"].width = 12  # Betyg Fredrik
ws_data.column_dimensions["L"].width = 12  # Betyg Ulrika
ws_data.column_dimensions["M"].width = 35  # Goodreads URL
ws_data.column_dimensions["N"].width = 35  # Notat

ws_data.merge_cells("A1:N1")
ws_data["A1"] = "📖 Bokdata — importerad från CSV (Notion-export)"
ws_data["A1"].font = Font(name="Arial", bold=True, size=12, color=C_WHITE)
ws_data["A1"].fill = fill(C_HEADER_DARK)
ws_data["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws_data.row_dimensions[1].height = 28

headers = ["#", "Titel", "Författare", "Hyllplats", "Språk", "Ägare",
           "År", "ISBN", "Förlag", "Format",
           "Betyg Fredrik", "Betyg Ulrika", "Goodreads", "Notat"]

for col, h in enumerate(headers, 1):
    c = ws_data.cell(row=2, column=col, value=h)
    c.font = Font(name="Arial", bold=True, size=9, color=C_WHITE)
    c.fill = fill(C_SHELF)
    c.alignment = Alignment(horizontal="center", wrap_text=True)
    c.border = border_thin()
ws_data.row_dimensions[2].height = 22

# Placeholder rows (3–7) to show format
placeholder = [
    [1, "[Importera CSV — se Instructions]", "", "", "", "", "", "", "", "", "", "", "", ""],
    [2, "← Kopiera data från CSV hit", "", "", "", "", "", "", "", "", "", "", "", ""],
]

for ri, row_data in enumerate(placeholder):
    for ci, val in enumerate(row_data, 1):
        c = ws_data.cell(row=3+ri, column=ci, value=val)
        c.font = Font(name="Arial", italic=True, size=9, color=C_INK_SOFT)
        c.fill = fill(C_CREAM_DARK if ri % 2 == 0 else C_CREAM)
        c.border = border_thin()
    ws_data.row_dimensions[3+ri].height = 18

ws_data.freeze_panes = "B3"

# ════════════════════════════════════════════════════════════
# FLIK 4 – Analysis
# ════════════════════════════════════════════════════════════
ws_anal = wb.create_sheet("Analysis")
ws_anal.sheet_properties.tabColor = C_HEADER_MED

ws_anal.column_dimensions["A"].width = 3
ws_anal.column_dimensions["B"].width = 28
ws_anal.column_dimensions["C"].width = 18
ws_anal.column_dimensions["D"].width = 18
ws_anal.column_dimensions["E"].width = 18
ws_anal.column_dimensions["F"].width = 18

ws_anal.merge_cells("A1:F1")
ws_anal["A1"] = "📊 Analys & Statistik"
ws_anal["A1"].font = Font(name="Arial", bold=True, size=14, color=C_WHITE)
ws_anal["A1"].fill = fill(C_HEADER_DARK)
ws_anal["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws_anal.row_dimensions[1].height = 32

ws_anal.merge_cells("A2:F2")
ws_anal["A2"] = '=CONCATENATE("Hyllägare: ", Settings!C4, "  |  Export: ", Settings!C5)'
ws_anal["A2"].font = Font(name="Arial", italic=True, size=9, color=C_INK_SOFT)
ws_anal["A2"].fill = fill(C_CREAM_DARK)
ws_anal["A2"].alignment = Alignment(horizontal="center")
ws_anal.row_dimensions[2].height = 18

# Summary stats
stats_header_row = 4
ws_anal.merge_cells(f"A{stats_header_row}:F{stats_header_row}")
ws_anal[f"A{stats_header_row}"] = "  SAMMANFATTNING"
ws_anal[f"A{stats_header_row}"].font = Font(name="Arial", bold=True, size=9, color=C_WHITE)
ws_anal[f"A{stats_header_row}"].fill = fill(C_HEADER_MED)
ws_anal.row_dimensions[stats_header_row].height = 20

summary = [
    ("Totalt antal böcker", '=COUNTA(Data!B3:B10000)-COUNTIF(Data!B3:B10000,"")'),
    ("Antal svenska böcker", '=COUNTIF(Data!E3:E10000,"sv")'),
    ("Antal engelska böcker", '=COUNTIF(Data!E3:E10000,"en")'),
    ("Antal hyllor", '=IFERROR(SUMPRODUCT(1/COUNTIF(Data!D3:D10000,Data!D3:D10000)*(Data!D3:D10000<>"")*(Data!B3:B10000<>"")),0)'),
    ("Böcker med betyg (Fredrik)", '=COUNTA(Data!K3:K10000)-COUNTIF(Data!K3:K10000,"")'),
    ("Böcker med betyg (Ulrika)", '=COUNTA(Data!L3:L10000)-COUNTIF(Data!L3:L10000,"")'),
    ("Medel­betyg Fredrik", '=IFERROR(AVERAGEIF(Data!K3:K10000,">0"),"-")'),
    ("Medel­betyg Ulrika", '=IFERROR(AVERAGEIF(Data!L3:L10000,">0"),"-")'),
]

for i, (label, formula) in enumerate(summary):
    r = stats_header_row + 1 + i
    ws_anal[f"B{r}"] = label
    ws_anal[f"B{r}"].font = Font(name="Arial", size=9, color=C_INK)
    ws_anal[f"B{r}"].fill = fill(C_CREAM if i % 2 == 0 else C_CREAM_DARK)
    ws_anal[f"B{r}"].border = border_thin()

    ws_anal[f"C{r}"] = formula
    ws_anal[f"C{r}"].font = Font(name="Arial", bold=True, size=10, color=C_FORMULA)
    ws_anal[f"C{r}"].fill = fill(C_CREAM if i % 2 == 0 else C_CREAM_DARK)
    ws_anal[f"C{r}"].border = border_thin()
    ws_anal[f"C{r}"].alignment = Alignment(horizontal="center")
    ws_anal.row_dimensions[r].height = 20

# Hyllor-räkning header
shelf_start = stats_header_row + len(summary) + 3
ws_anal.merge_cells(f"A{shelf_start}:F{shelf_start}")
ws_anal[f"A{shelf_start}"] = "  BÖCKER PER HYLLA"
ws_anal[f"A{shelf_start}"].font = Font(name="Arial", bold=True, size=9, color=C_WHITE)
ws_anal[f"A{shelf_start}"].fill = fill(C_HEADER_MED)
ws_anal.row_dimensions[shelf_start].height = 20

# Header
for col, txt in [(2, "Hylla"), (3, "Antal böcker"), (4, "Varav svenska"), (5, "Varav engelska")]:
    c = ws_anal.cell(row=shelf_start+1, column=col, value=txt)
    c.font = Font(name="Arial", bold=True, size=9, color=C_WHITE)
    c.fill = fill(C_SHELF)
    c.alignment = Alignment(horizontal="center")
    c.border = border_thin()

shelves = ["A1","A2","A3","A4","A5","A6",
           "B1","B2","B3","B4","B5","B6",
           "C1","C2","C3","C4","C5","C6",
           "Kök1","Kök2","W1","W2","W3"]

for i, shelf in enumerate(shelves):
    r = shelf_start + 2 + i
    bg = C_CREAM if i % 2 == 0 else C_CREAM_DARK

    ws_anal[f"B{r}"] = shelf
    ws_anal[f"B{r}"].font = Font(name="Arial", bold=True, size=9, color=C_SHELF)
    ws_anal[f"B{r}"].fill = fill(bg)
    ws_anal[f"B{r}"].border = border_thin()
    ws_anal[f"B{r}"].alignment = Alignment(horizontal="center")

    ws_anal[f"C{r}"] = f'=COUNTIF(Data!$D$3:$D$10000,"{shelf}")'
    ws_anal[f"C{r}"].font = Font(name="Arial", size=9, color=C_FORMULA)
    ws_anal[f"C{r}"].fill = fill(bg)
    ws_anal[f"C{r}"].border = border_thin()
    ws_anal[f"C{r}"].alignment = Alignment(horizontal="center")

    ws_anal[f"D{r}"] = f'=COUNTIFS(Data!$D$3:$D$10000,"{shelf}",Data!$E$3:$E$10000,"sv")'
    ws_anal[f"D{r}"].font = Font(name="Arial", size=9, color=C_FORMULA)
    ws_anal[f"D{r}"].fill = fill(bg)
    ws_anal[f"D{r}"].border = border_thin()
    ws_anal[f"D{r}"].alignment = Alignment(horizontal="center")

    ws_anal[f"E{r}"] = f'=COUNTIFS(Data!$D$3:$D$10000,"{shelf}",Data!$E$3:$E$10000,"en")'
    ws_anal[f"E{r}"].font = Font(name="Arial", size=9, color=C_FORMULA)
    ws_anal[f"E{r}"].fill = fill(bg)
    ws_anal[f"E{r}"].border = border_thin()
    ws_anal[f"E{r}"].alignment = Alignment(horizontal="center")

    ws_anal.row_dimensions[r].height = 18

# Totals row
total_r = shelf_start + 2 + len(shelves)
ws_anal[f"B{total_r}"] = "TOTALT"
ws_anal[f"B{total_r}"].font = Font(name="Arial", bold=True, size=9, color=C_WHITE)
ws_anal[f"B{total_r}"].fill = fill(C_HEADER_DARK)
ws_anal[f"B{total_r}"].border = border_thin()
ws_anal[f"B{total_r}"].alignment = Alignment(horizontal="center")

for col_letter in ["C", "D", "E"]:
    col_start = shelf_start + 2
    col_end = shelf_start + 1 + len(shelves)
    ws_anal[f"{col_letter}{total_r}"] = f"=SUM({col_letter}{col_start}:{col_letter}{col_end})"
    ws_anal[f"{col_letter}{total_r}"].font = Font(name="Arial", bold=True, size=9, color=C_GOLD_LIGHT)
    ws_anal[f"{col_letter}{total_r}"].fill = fill(C_HEADER_DARK)
    ws_anal[f"{col_letter}{total_r}"].border = border_thin()
    ws_anal[f"{col_letter}{total_r}"].alignment = Alignment(horizontal="center")

ws_anal.row_dimensions[total_r].height = 22

# ════════════════════════════════════════════════════════════
# FLIK 5 – File History
# ════════════════════════════════════════════════════════════
ws_hist = wb.create_sheet("File History")
ws_hist.sheet_properties.tabColor = "8B6914"

ws_hist.column_dimensions["A"].width = 10
ws_hist.column_dimensions["B"].width = 16
ws_hist.column_dimensions["C"].width = 20
ws_hist.column_dimensions["D"].width = 55

ws_hist.merge_cells("A1:D1")
ws_hist["A1"] = "📋 Versionshistorik"
ws_hist["A1"].font = Font(name="Arial", bold=True, size=13, color=C_WHITE)
ws_hist["A1"].fill = fill(C_HEADER_DARK)
ws_hist["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws_hist.row_dimensions[1].height = 28

for col, txt in [(1, "Version"), (2, "Datum"), (3, "Skapad av"), (4, "Ändringar")]:
    c = ws_hist.cell(row=2, column=col, value=txt)
    c.font = Font(name="Arial", bold=True, size=9, color=C_WHITE)
    c.fill = fill(C_SHELF)
    c.alignment = Alignment(horizontal="center")
    c.border = border_thin()
ws_hist.row_dimensions[2].height = 20

history = [
    ("v1.0", datetime.date.today().strftime("%Y-%m-%d"), "Clio / Fredrik Arvas",
     "Initial version. 5-fliksstruktur. CSV-baserad dataimport. Analysis med hyllräkning och betygsstatistik. AIAB-färgpalett."),
]

for i, (ver, date, author, changes) in enumerate(history):
    r = 3 + i
    bg = C_CREAM if i % 2 == 0 else C_CREAM_DARK
    for col, val in enumerate([ver, date, author, changes], 1):
        c = ws_hist.cell(row=r, column=col, value=val)
        c.font = Font(name="Arial", size=9, color=C_INK if col > 1 else C_INPUT_VAL,
                      bold=(col == 1))
        c.fill = fill(bg)
        c.border = border_thin()
        c.alignment = Alignment(wrap_text=(col == 4), vertical="top")
    ws_hist.row_dimensions[r].height = 40

# ── Spara ────────────────────────────────────────────────────
output_path = "/home/claude/Arvas_Bibliotek_Mall_v1.0.xlsx"
wb.save(output_path)
print(f"Sparad: {output_path}")
