#!/usr/bin/env python3
"""
Arvas Familjebibliotek — Sprint 3: Preprocessing av råfiler
Skannar import/queue/ efter GoodReads Excel och Storytel CSV,
transformerar till standardformat (bokregister_*.csv + betyg_*.csv).

Stöder:
  - GoodReads Excel (copy-paste från webben): *.xlsx
  - Storytel CSV-export (lyssningshistorik): *listening-history*.csv

Användning:
  python prepare_import.py --dry-run            # visa vad som hittas
  python prepare_import.py                      # transformera och skriv output
  python prepare_import.py --min-progress 85    # Storytel: min % för "läst"
  python prepare_import.py --skip-isbn-lookup   # hoppa över OpenLibrary-lookup
"""

import argparse
import csv
import io
import json
import logging
import os
import re
import shutil
import sys
import time
import urllib.request
import urllib.error
import urllib.parse
from datetime import datetime, timedelta
from pathlib import Path

# Tvinga UTF-8 i stdout/stderr för att svenska tecken ska synas rätt i loggen
if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

# ─── Konfig ────────────────────────────────────────────────────────────────────
HERE         = Path(__file__).parent
QUEUE_DIR    = HERE / "import" / "queue"
IMPORTED_DIR = HERE / "import" / "imported"
LOG_FILE     = HERE / "prepare_import.log"


def _load_google_api_key() -> str:
    env_file = HERE / ".env"
    if env_file.exists():
        for line in env_file.read_text(encoding="utf-8").splitlines():
            if line.strip().startswith("GOOGLE_API_KEY="):
                return line.split("=", 1)[1].strip()
    return os.environ.get("GOOGLE_API_KEY", "")

GOODREADS_RATING_MAP = {
    "it was amazing":   5,
    "really liked it":  4,
    "liked it":         3,
    "it was ok":        2,
    "did not like it":  1,
}

STORYTEL_FORMAT_MAP = {
    "Ljudbok": "audio",
    "E-bok":   "ebook",
}

# Person-namn härleds från filnamn
PERSON_PATTERNS = {
    r"(?i)alice":   "Alice",
    r"(?i)johan":   "Johan",
    r"(?i)ulrika":  "Ulrika",
    r"(?i)fredrik": "Fredrik",
}


# ─── Logging ────────────────────────────────────────────────────────────────────
def setup_logging():
    logger = logging.getLogger()
    logger.setLevel(logging.DEBUG)
    fmt = logging.Formatter("%(asctime)s %(levelname)-8s %(message)s", "%H:%M:%S")
    ch = logging.StreamHandler()
    ch.setLevel(logging.INFO)
    ch.setFormatter(fmt)
    fh = logging.FileHandler(LOG_FILE, encoding="utf-8")
    fh.setLevel(logging.DEBUG)
    fh.setFormatter(fmt)
    logger.addHandler(ch)
    logger.addHandler(fh)


# ─── Hjälpfunktioner ───────────────────────────────────────────────────────────
def detect_person(filename: str) -> str:
    """Härleda person från filnamn."""
    for pattern, name in PERSON_PATTERNS.items():
        if re.search(pattern, filename):
            return name
    return "Okänd"


def clean_title(title: str) -> str:
    """Rensa seriesuffix och whitespace från titel."""
    if not title:
        return ""
    title = title.replace("\xa0", " ").strip()
    # Ta bort seriesuffix: "Titel (Serie, #3)" → "Titel"
    title = re.sub(r'\s*\([^)]*#\d+[^)]*\)\s*$', '', title)
    # Ta bort trailing whitespace/nbsp
    return title.strip()


def flip_author(author: str) -> tuple:
    """Flippa "Efternamn, Förnamn *" → ("Förnamn Efternamn", "Efternamn", "Förnamn").
    Returnerar (full_name, efternamn, förnamn)."""
    if not author:
        return ("", "", "")
    author = author.replace("\xa0", " ").strip()
    # Ta bort trailing * (GoodReads-markering)
    author = re.sub(r'\s*\*\s*$', '', author)
    author = author.strip()

    if "," in author:
        parts = author.split(",", 1)
        enamn = parts[0].strip()
        fnamn = parts[1].strip() if len(parts) > 1 else ""
        full = f"{fnamn} {enamn}".strip()
        return (full, enamn, fnamn)
    else:
        # Redan i "Förnamn Efternamn"-format
        parts = author.rsplit(" ", 1)
        fnamn = parts[0] if len(parts) > 1 else author
        enamn = parts[1] if len(parts) > 1 else ""
        return (author, enamn, fnamn)


def parse_goodreads_date(date_str: str) -> str:
    """Parsea GoodReads-datum ("Mar 17, 2026" | "not set") → ISO 8601 eller tom."""
    if not date_str or date_str.strip().lower() in ("not set", "none", ""):
        return ""
    date_str = date_str.replace("\xa0", " ").strip()
    try:
        dt = datetime.strptime(date_str, "%b %d, %Y")
        return dt.strftime("%Y-%m-%d")
    except ValueError:
        try:
            dt = datetime.strptime(date_str, "%B %d, %Y")
            return dt.strftime("%Y-%m-%d")
        except ValueError:
            logging.debug("Kunde inte parsea datum: '%s'", date_str)
            return ""


def parse_storytel_datetime(dt_str: str) -> str:
    """Parsea Storytel ISO-datetime → datum."""
    if not dt_str:
        return ""
    try:
        # "2026-02-13T06:46:19Z"
        dt = datetime.fromisoformat(dt_str.replace("Z", "+00:00"))
        return dt.strftime("%Y-%m-%d")
    except ValueError:
        return dt_str[:10] if len(dt_str) >= 10 else ""


def parse_storytel_full_datetime(dt_str: str):
    """Parsea Storytel ISO-datetime → datetime-objekt."""
    if not dt_str:
        return None
    try:
        return datetime.fromisoformat(dt_str.replace("Z", "+00:00"))
    except ValueError:
        return None


# ─── ISBN-lookup (OpenLibrary → Google Books → Libris) ─────────────────────────
def _api_get(url: str, timeout: int = 5) -> dict:
    """Hämtar JSON från URL, returnerar {} vid fel."""
    req = urllib.request.Request(url, headers={"User-Agent": "ArvasLibrary/1.0"})
    try:
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            return json.loads(resp.read().decode("utf-8"))
    except Exception:
        return {}


def _lookup_openlib(isbn: str) -> str:
    data = _api_get(f"https://openlibrary.org/isbn/{isbn}.json")
    author_keys = data.get("authors", [])
    if not author_keys:
        works = data.get("works", [])
        if works:
            work_data = _api_get(f"https://openlibrary.org{works[0].get('key','')}.json")
            author_keys = [a.get("author", a) for a in work_data.get("authors", [])]
    if author_keys:
        key = author_keys[0].get("key", "")
        if key:
            return _api_get(f"https://openlibrary.org{key}.json").get("name", "")
    return ""


def _lookup_google(isbn: str) -> str:
    key = _load_google_api_key()
    if not key:
        return ""
    data = _api_get(
        f"https://www.googleapis.com/books/v1/volumes?q=isbn:{isbn}&key={key}"
    )
    items = data.get("items", [])
    if items:
        authors = items[0].get("volumeInfo", {}).get("authors", [])
        return authors[0] if authors else ""
    return ""


def _lookup_libris(isbn: str) -> str:
    data = _api_get(
        f"https://libris.kb.se/xsearch?query=isbn:{isbn}&format=json&n=1"
    )
    hits = data.get("xsearch", {}).get("list", [])
    if hits:
        creator = hits[0].get("creator", "")
        # Libris format: "Efternamn, Förnamn, 1974-" — strippa födelseår
        creator = re.sub(r",?\s*\d{4}-?\s*$", "", creator).strip()
        return creator
    return ""


def lookup_author_by_isbn(isbn: str) -> str:
    """Slå upp författare via OpenLibrary → Google Books → Libris."""
    if not isbn:
        return ""
    for fn, name in [(_lookup_openlib, "OpenLibrary"),
                     (_lookup_google,  "Google Books"),
                     (_lookup_libris,  "Libris")]:
        try:
            result = fn(isbn)
            if result:
                logging.debug("  ISBN %s → %s (%s)", isbn, result, name)
                return result
        except Exception as e:
            logging.debug("  %s fel för %s: %s", name, isbn, e)
    return ""


# ─── GoodReads-parser ──────────────────────────────────────────────────────────
def parse_goodreads_excel(filepath: Path) -> tuple:
    """Parsea GoodReads Excel (copy-paste från webben).
    Returnerar (bokregister_rows, betyg_rows)."""
    import openpyxl

    person = detect_person(filepath.stem)
    logging.info("  Person: %s (härlett från filnamn)", person)

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    bokregister = []
    betyg = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        logging.info("  Flik: %s", sheet_name)

        for row in ws.iter_rows(values_only=True):
            cells = [str(c).strip() if c is not None else "" for c in row]

            # Hitta #VALUE!-ankare
            try:
                anchor = cells.index("#VALUE!")
            except ValueError:
                continue  # Inte en bokrad

            # Extrahera med offset från ankare
            title_raw     = cells[anchor + 1] if anchor + 1 < len(cells) else ""
            author_raw    = cells[anchor + 2] if anchor + 2 < len(cells) else ""
            # anchor+3 = avg_rating (skippa)
            rating_text   = cells[anchor + 4] if anchor + 4 < len(cells) else ""
            # anchor+5 = star display (skippa)
            date_read_raw = cells[anchor + 6] if anchor + 6 < len(cells) else ""
            date_added_raw = cells[anchor + 7] if anchor + 7 < len(cells) else ""

            if not title_raw or title_raw.lower() == "none":
                continue

            title = clean_title(title_raw)
            full_author, enamn, fnamn = flip_author(author_raw)

            # Betyg: mappa textbetyg
            betyg_val = ""
            rating_lower = rating_text.lower().strip()
            if rating_lower in GOODREADS_RATING_MAP:
                betyg_val = str(GOODREADS_RATING_MAP[rating_lower])

            date_read = parse_goodreads_date(date_read_raw)
            date_added = parse_goodreads_date(date_added_raw)

            # Bokregister-rad
            bokregister.append({
                "Titel": title,
                "Författare": full_author,
                "Författare_Enamn": enamn,
                "Författare_Fnamn": fnamn,
                "Hyllplats": "",
                "Språk": "",
                "Format": "",
                "Hus": "",
            })

            # Betyg-rad
            betyg.append({
                "Titel": title,
                "Författare": full_author,
                "Person": person,
                "Betyg": betyg_val,
                "Källa": "Goodreads",
                "Datum läst": date_read,
                "Datum tillagt": date_added,
            })

    wb.close()
    return bokregister, betyg


# ─── Storytel-parser ───────────────────────────────────────────────────────────
def parse_storytel_csv(filepath: Path, min_progress: int = 85,
                       skip_isbn_lookup: bool = False) -> tuple:
    """Parsea Storytel lyssningshistorik CSV.
    Returnerar (bokregister_rows, betyg_rows)."""
    person = detect_person(filepath.stem)
    logging.info("  Person: %s (härlett från filnamn)", person)

    rows = []
    with open(filepath, encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            rows.append({k.strip(): v.strip() if v else "" for k, v in row.items()})

    logging.info("  %d rader laddade", len(rows))

    bokregister = []
    betyg = []
    skipped = 0
    isbn_cache = {}
    lookup_count = 0

    for i, row in enumerate(rows, 1):
        titel = row.get("Titel", "").strip()
        isbn = row.get("Bokens ISBN", "").strip()
        position_str = row.get("Bokposition", "").replace("%", "").strip()
        format_raw = row.get("Format", "")
        start_str = row.get("Startdatum", "")
        bookmark_str = row.get("Senaste bokmärket skapades", "")

        if not titel:
            continue

        # Klassificering: läst vs avbruten
        try:
            position = int(position_str)
        except ValueError:
            position = 0

        start_dt = parse_storytel_full_datetime(start_str)
        bookmark_dt = parse_storytel_full_datetime(bookmark_str)

        is_read = False
        if position >= min_progress:
            if start_dt and bookmark_dt:
                duration = bookmark_dt - start_dt
                if duration > timedelta(hours=1):
                    is_read = True
            else:
                # Om vi inte kan beräkna tid, lita på position
                is_read = True

        if not is_read:
            skipped += 1
            logging.debug("  [%d] Avbruten: %s (%d%%)", i, titel[:40], position)
            continue

        # Format
        fmt = STORYTEL_FORMAT_MAP.get(format_raw, "audio")

        # Författare via ISBN-lookup
        author = ""
        if isbn and not skip_isbn_lookup:
            if isbn in isbn_cache:
                author = isbn_cache[isbn]
            else:
                author = lookup_author_by_isbn(isbn)
                isbn_cache[isbn] = author
                lookup_count += 1
                logging.info("  [ISBN %d] %s → %s", lookup_count, isbn, author or "—")
                time.sleep(0.3)

        full_author, enamn, fnamn = flip_author(author) if author else ("", "", "")

        # Datum
        datum_last = parse_storytel_datetime(start_str)

        bokregister.append({
            "Titel": titel,
            "Författare": full_author,
            "Författare_Enamn": enamn,
            "Författare_Fnamn": fnamn,
            "Hyllplats": "",
            "Språk": "",
            "Format": fmt,
            "Hus": "",
            "ISBN": isbn,
        })

        betyg.append({
            "Titel": titel,
            "Författare": full_author,
            "Person": person,
            "Betyg": "",
            "Källa": "Storytel",
            "Datum läst": datum_last,
            "Datum tillagt": datum_last,
        })

    # Deduplicera: behåll senaste lyssning per titel
    seen_titles = {}
    deduped_bok = []
    deduped_bet = []
    for bok_row, bet_row in zip(bokregister, betyg):
        titel = bok_row["Titel"]
        if titel in seen_titles:
            # Behåll den med senaste datum
            old_idx = seen_titles[titel]
            old_date = deduped_bet[old_idx].get("Datum läst", "")
            new_date = bet_row.get("Datum läst", "")
            if new_date > old_date:
                deduped_bok[old_idx] = bok_row
                deduped_bet[old_idx] = bet_row
        else:
            seen_titles[titel] = len(deduped_bok)
            deduped_bok.append(bok_row)
            deduped_bet.append(bet_row)

    dupes = len(bokregister) - len(deduped_bok)
    if dupes:
        logging.info("  Dubbletter borttagna: %d", dupes)

    logging.info("  Lästa: %d (unika), Avbrutna: %d", len(deduped_bok), skipped)
    return deduped_bok, deduped_bet


# ─── CSV-output ─────────────────────────────────────────────────────────────────
BOKREGISTER_FIELDS = ["Titel", "Författare", "Författare_Enamn", "Författare_Fnamn",
                      "Hyllplats", "Språk", "Format", "Hus", "ISBN"]
BETYG_FIELDS = ["Titel", "Författare", "Person", "Betyg", "Källa",
                "Datum läst", "Datum tillagt"]


def write_csv(rows: list, fields: list, filepath: Path):
    """Skriv CSV med semikolon-separator och UTF-8 BOM."""
    with open(filepath, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fields, delimiter=";",
                                extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


# ─── Källdetektering ───────────────────────────────────────────────────────────
def detect_source(filepath: Path) -> str:
    """Detekterar källa baserat på filnamn och innehåll."""
    name = filepath.name.lower()

    # Redan normaliserade → hoppa över
    if name.startswith("bokregister") or name.startswith("betyg"):
        return "normalized"

    # GoodReads Excel
    if filepath.suffix.lower() == ".xlsx":
        return "goodreads"

    # Storytel CSV
    if "listening-history" in name or "storytel" in name:
        return "storytel"

    # Okänd CSV — försök detektera från kolumner
    if filepath.suffix.lower() == ".csv":
        try:
            with open(filepath, encoding="utf-8-sig") as f:
                reader = csv.DictReader(f)
                cols = set(reader.fieldnames or [])
            if "Bokens ISBN" in cols or "Bokposition" in cols:
                return "storytel"
        except Exception:
            pass

    return "unknown"


# ─── Main ────────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="Preprocessa råfiler till importformat")
    parser.add_argument("--dry-run", action="store_true",
                        help="Visa vad som hittas utan att skriva output")
    parser.add_argument("--min-progress", type=int, default=85,
                        help="Storytel: min %% lyssnande för 'läst' (default: 85)")
    parser.add_argument("--skip-isbn-lookup", action="store_true",
                        help="Hoppa över OpenLibrary ISBN-lookup för Storytel")
    parser.add_argument("--source-dir", type=Path, default=QUEUE_DIR,
                        help="Alternativ indata-mapp")
    args = parser.parse_args()

    setup_logging()

    source_dir = args.source_dir
    if not source_dir.exists():
        logging.error("Mappen %s finns inte", source_dir)
        return

    # Hitta filer
    files = sorted(source_dir.iterdir())
    files = [f for f in files if f.is_file() and f.suffix.lower() in (".xlsx", ".csv")]

    if not files:
        logging.error("Inga .xlsx/.csv-filer i %s", source_dir)
        return

    logging.info("Hittade %d fil(er) i %s", len(files), source_dir)

    ts = datetime.now().strftime("%Y%m%d")
    total_books = 0
    total_ratings = 0

    for filepath in files:
        source = detect_source(filepath)
        logging.info("\n── %s (%s) ──", filepath.name, source)

        if source == "normalized":
            logging.info("  Redan normaliserad — hoppar över")
            continue

        if source == "unknown":
            logging.warning("  Okänt format — hoppar över")
            continue

        try:
            if source == "goodreads":
                bok_rows, bet_rows = parse_goodreads_excel(filepath)
            elif source == "storytel":
                bok_rows, bet_rows = parse_storytel_csv(
                    filepath,
                    min_progress=args.min_progress,
                    skip_isbn_lookup=args.skip_isbn_lookup or args.dry_run,
                )
            else:
                continue
        except Exception as e:
            logging.error("  Fel vid parsning: %s", e)
            continue

        person = detect_person(filepath.stem)
        logging.info("  Böcker: %d, Betyg: %d", len(bok_rows), len(bet_rows))
        total_books += len(bok_rows)
        total_ratings += len(bet_rows)

        if args.dry_run:
            # Visa sample
            for row in bok_rows[:3]:
                logging.info("  BOK: %s | %s", row["Titel"][:40], row["Författare"][:25])
            for row in bet_rows[:3]:
                logging.info("  BET: %s | %s | betyg=%s | %s",
                             row.get("Titel", "")[:30], row["Person"],
                             row["Betyg"] or "-", row["Källa"])
            if len(bok_rows) > 3:
                logging.info("  ... och %d till", len(bok_rows) - 3)
            continue

        # Skriv output
        bok_path = source_dir / f"bokregister_{source}_{person.lower()}_{ts}.csv"
        bet_path = source_dir / f"betyg_{source}_{person.lower()}_{ts}.csv"

        write_csv(bok_rows, BOKREGISTER_FIELDS, bok_path)
        logging.info("  → %s (%d rader)", bok_path.name, len(bok_rows))

        write_csv(bet_rows, BETYG_FIELDS, bet_path)
        logging.info("  → %s (%d rader)", bet_path.name, len(bet_rows))

        # Flytta original till imported
        dest = IMPORTED_DIR / f"{filepath.stem}_{ts}{filepath.suffix}"
        shutil.move(str(filepath), str(dest))
        logging.info("  Original → %s", dest.name)

    logging.info("\n── Sammanfattning ──")
    logging.info("  Böcker totalt:  %d", total_books)
    logging.info("  Betyg totalt:   %d", total_ratings)

    if args.dry_run:
        logging.info("\n  DRY-RUN — inget skrivet, original ej flyttade.")
    else:
        logging.info("\n  Output i %s — kör import_books.py + import_lasningar.py härnäst.", source_dir)


if __name__ == "__main__":
    main()
