"""
watchlist/send_invitation.py — Skickar inbjudningsmail med personanpassad bevakningslista-mall

Bilagan är en .xlsx-fil förformaterad med mottagarens e-post som filnamn och deras
eget namn som exempelrad, redo att öppnas i Excel och fyllas i.

Mottagaren svarar med [clio-obit] i ämnesraden — auto-import via clio-agent-mail (Sprint 3).

Körning:
    python send_invitation.py --to-name "Ulrika Arvas" --to-email ulrika@arvas.se
    python send_invitation.py --to-name "Ulrika Arvas" --to-email ulrika@arvas.se --dry-run
"""

from __future__ import annotations

import argparse
import io
import os
import sys
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))
from notifier import _load_config, _create_smtp


def _make_xlsx(to_name: str, to_email: str) -> bytes:
    """Skapar en personanpassad Excel-mall och returnerar den som bytes."""
    parts = to_name.strip().split(None, 1)
    fornamn   = parts[0] if parts else to_name
    efternamn = parts[1] if len(parts) > 1 else ""

    wb = openpyxl.Workbook()

    # ── Blad 1: Bevakningslista ───────────────────────────────────────────────
    ws = wb.active
    ws.title = "Bevakningslista"

    # Stilar
    header_font    = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill    = PatternFill("solid", fgColor="2C3E50")
    info_font      = Font(name="Calibri", italic=True, color="7F8C8D", size=10)
    info_fill      = PatternFill("solid", fgColor="F8F9FA")
    example_font   = Font(name="Calibri", color="2C3E50", size=11)
    example_fill   = PatternFill("solid", fgColor="EAF4FB")
    thin_border    = Border(
        bottom=Side(style="thin", color="D5D8DC"),
    )

    # Rad 1: instruktionstext (sammanfogad)
    ws.merge_cells("A1:F1")
    ws["A1"] = f"Bevakningslista för {to_email} — fyll i och svara på mailet med [clio-obit] i ämnesraden."
    ws["A1"].font = info_font
    ws["A1"].fill = info_fill
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 28

    # Rad 2: kolumnrubriker
    headers = ["efternamn", "fornamn", "fodelsear", "hemort", "prioritet", "kalla"]
    header_help = [
        "Obligatoriskt",
        "Obligatoriskt",
        "t.ex. 1948 (lämna tomt om okänt)",
        "t.ex. Haninge (lämna tomt om okänt)",
        "viktig / normal / bra_att_veta",
        "manuell",
    ]
    for col, (h, _) in enumerate(zip(headers, header_help), start=1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    ws.row_dimensions[2].height = 22

    # Rad 3: exempelrad med mottagarens eget namn
    example = [
        efternamn if efternamn else "Efternamn",
        fornamn,
        "",
        "",
        "viktig",
        "manuell",
    ]
    for col, val in enumerate(example, start=1):
        cell = ws.cell(row=3, column=col, value=val)
        cell.font = example_font
        cell.fill = example_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[3].height = 20

    # Rader 4–13: tomma datarader (lätt grå alternering)
    empty_fill = PatternFill("solid", fgColor="FFFFFF")
    alt_fill   = PatternFill("solid", fgColor="F2F3F4")
    for row in range(4, 14):
        fill = alt_fill if row % 2 == 0 else empty_fill
        for col in range(1, 7):
            cell = ws.cell(row=row, column=col, value="")
            cell.fill = fill
            cell.font = Font(name="Calibri", size=11)
        ws.row_dimensions[row].height = 20

    # Kolumnbredder
    from openpyxl.utils import get_column_letter
    col_widths = [18, 16, 14, 16, 22, 12]
    for col, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width

    # Frys rubrikrad
    ws.freeze_panes = "A3"

    # ── Blad 2: Instruktioner ─────────────────────────────────────────────────
    wi = wb.create_sheet("Instruktioner")
    wi.column_dimensions["A"].width = 70

    title_font = Font(name="Calibri", bold=True, size=14, color="2C3E50")
    body_font  = Font(name="Calibri", size=11, color="2C3E50")
    code_font  = Font(name="Courier New", size=10, color="7D3C98")

    rows = [
        (f"Dödsannonsbevakning — {to_email}", title_font, 24),
        ("", body_font, 8),
        ("Så här gör du:", Font(name="Calibri", bold=True, size=11, color="2C3E50"), 18),
        ("1. Fyll i de personer du vill bevaka i bladet 'Bevakningslista'.", body_font, 16),
        ("   Rad 3 (blå) är din exempelrad — ändra eller ta bort den.", body_font, 16),
        ("2. Spara filen.", body_font, 16),
        ("3. Svara på mailet du fick med den här filen bifogad.", body_font, 16),
        ("   Behåll '[clio-obit]' i ämnesraden — då importeras listan automatiskt.", body_font, 16),
        ("", body_font, 8),
        ("Prioritetsnivåer:", Font(name="Calibri", bold=True, size=11, color="2C3E50"), 18),
        ("  viktig       →  du får ett mail direkt om personen dyker upp i en dödsannons", body_font, 16),
        ("  normal       →  samlas i en daglig sammanfattning", body_font, 16),
        ("  bra_att_veta →  daglig sammanfattning, lägre prioritet", body_font, 16),
        ("", body_font, 8),
        ("Obligatoriska fält: efternamn, fornamn", Font(name="Calibri", bold=True, size=10, color="E74C3C"), 16),
        ("Valfria fält: fodelsear (förbättrar träffsäkerheten), hemort", body_font, 14),
        ("", body_font, 8),
        ("Frågor? Svara på mailet.", Font(name="Calibri", italic=True, size=10, color="7F8C8D"), 14),
    ]

    for r, (text, font, height) in enumerate(rows, start=1):
        cell = wi.cell(row=r, column=1, value=text)
        cell.font = font
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        wi.row_dimensions[r].height = height

    # Spara till bytes
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


_BODY = """\
Hej {fornamn},

Du är nu inbjuden till dödsannonsbevakningen — ett system som automatiskt
söker igenom svenska dödsannonser och meddelar dig om någon på din lista dyker upp.

Så här gör du:

  1. Öppna den bifogade Excel-filen {filename}.
  2. Fyll i de personer du vill bevaka i bladet "Bevakningslista".
     Rad 3 (blå) är din exempelrad — ändra eller ta bort den.
  3. Spara filen.
  4. Svara på det här mailet med filen bifogad.
     Viktigt: behåll "[clio-obit]" i ämnesraden så importeras den automatiskt.

Instruktioner och prioritetsnivåer finns i bladet "Instruktioner" i filen.

Hör av dig om du har frågor.

— clio-agent-obit
"""


def send_invitation(to_name: str, to_email: str, dry_run: bool = False) -> None:
    cfg        = _load_config()
    smtp_user  = cfg.get("smtp", {}).get("user", "")
    from_label = cfg.get("notify", {}).get("from_label", "clio-agent-obit")

    fornamn  = to_name.strip().split()[0]
    filename = f"{to_email}.xlsx"
    xlsx_bytes = _make_xlsx(to_name, to_email)

    subject = "[clio-obit] Dödsannonsbevakning — din bevakningslista"
    body    = _BODY.format(fornamn=fornamn, filename=filename)

    msg = MIMEMultipart()
    msg["Subject"] = subject
    msg["From"]    = f"{from_label} <{smtp_user}>"
    msg["To"]      = to_email
    msg["Reply-To"] = smtp_user

    msg.attach(MIMEText(body, "plain", "utf-8"))

    # Excel-bilaga
    part = MIMEBase(
        "application",
        "vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    part.set_payload(xlsx_bytes)
    encoders.encode_base64(part)
    part.add_header("Content-Disposition", "attachment", filename=filename)
    msg.attach(part)

    if dry_run:
        print(f"[dry-run] Till:    {to_name} <{to_email}>")
        print(f"[dry-run] Ämne:    {subject}")
        print(f"[dry-run] Bilaga:  {filename} ({len(xlsx_bytes):,} bytes)")
        print("\n[dry-run] Excel-filen skapades utan fel.")
        return

    with _create_smtp(cfg) as smtp:
        smtp.sendmail(smtp_user, to_email, msg.as_string())
    print(f"Inbjudan skickad till {to_name} <{to_email}>")


def parse_args(argv=None):
    p = argparse.ArgumentParser(description="Skicka inbjudningsmail med bevakningslista-mall")
    p.add_argument("--to-name",  required=True, help="Mottagarens fullständiga namn")
    p.add_argument("--to-email", required=True, help="Mottagarens e-postadress")
    p.add_argument("--dry-run",  action="store_true")
    return p.parse_args(argv)


def main(argv=None):
    args = parse_args(argv)
    send_invitation(args.to_name, args.to_email, dry_run=args.dry_run)


if __name__ == "__main__":
    main()
